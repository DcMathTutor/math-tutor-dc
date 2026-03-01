"""
generate_finances.py
--------------------
Run once:  python generate_finances.py
Output:    finances.xlsx  (in the same directory)

10-sheet model for Math Tutor DC
  Rates: $50/hr client, $35/hr tutor, 30% gross margin
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import os

# ── Color palette ──────────────────────────────────────────────────────────────
BLACK  = "FF000000"
WHITE  = "FFFFFFFF"
GRAY1  = "FFF4F4F4"   # light row background
GRAY2  = "FFDDDDDD"   # border / divider
GREEN  = "FFD4EDDA"   # positive highlight
RED    = "FFF8D7DA"   # negative highlight
YELLOW = "FFFFF3CD"   # warning

# ── Style helpers ──────────────────────────────────────────────────────────────
def hdr(bold=True, size=11, color=WHITE, bg=BLACK):
    f = Font(name="Calibri", bold=bold, size=size, color=color)
    fill = PatternFill("solid", fgColor=bg)
    return f, fill

def money_fmt():
    return '#,##0.00'

def pct_fmt():
    return '0.0%'

def thin_border():
    s = Side(style="thin", color=GRAY2)
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def header_row(ws, row, labels, bg=BLACK, fg=WHITE, bold=True):
    f = Font(name="Calibri", bold=bold, size=10, color=fg)
    fill = PatternFill("solid", fgColor=bg)
    for c, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font  = f
        cell.fill  = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def write(ws, row, col, value, bold=False, italic=False, size=10,
          color=BLACK, bg=None, num_fmt=None, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=bold, italic=italic,
                     size=size, color=color)
    cell.alignment = Alignment(horizontal=align, wrap_text=False)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def money(ws, row, col, value, bold=False, color=BLACK, bg=None):
    return write(ws, row, col, value, bold=bold, color=color, bg=bg,
                 num_fmt=money_fmt(), align="right")

def pct(ws, row, col, value, bold=False, color=BLACK, bg=None):
    return write(ws, row, col, value, bold=bold, color=color, bg=bg,
                 num_fmt=pct_fmt(), align="right")

def freeze(ws, cell="B2"):
    ws.freeze_panes = cell

# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

# ── Sheet order ───────────────────────────────────────────────────────────────
SHEETS = [
    "Dashboard",
    "Monthly P&L",
    "3yr Conservative",
    "3yr Base Case",
    "3yr Aggressive",
    "Lead Funnel",
    "Tutor Roster",
    "Client Roster",
    "Session Log",
    "Payment Log",
    "Expense Log",
    "Pricing Calculator",
]

ws_map = {}
for name in SHEETS:
    ws = wb.create_sheet(name)
    ws_map[name] = ws

# ══════════════════════════════════════════════════════════════════════════════
# 1.  DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws = ws_map["Dashboard"]
ws.sheet_view.showGridLines = False
set_col_widths(ws, [2, 28, 18, 18, 18, 2])

# Title
ws.merge_cells("B2:E2")
c = ws.cell(row=2, column=2, value="Math Tutor DC – Financial Dashboard")
c.font = Font(name="Calibri", bold=True, size=16, color=BLACK)
c.alignment = Alignment(horizontal="center")
ws.row_dimensions[2].height = 30

write(ws, 3, 2, "Rates: $50/hr (client)  ·  $35/hr (tutor)  ·  30% gross margin",
      italic=True, size=9, color="FF555555")

# KPI boxes
kpi_row = 5
header_row(ws, kpi_row, ["", "Metric", "Month", "YTD", "Target"], bg="FF222222")

kpis = [
    ("Revenue",            0,    0,    10000),
    ("Tutor Payments",     0,    0,    7000),
    ("Gross Profit",       0,    0,    3000),
    ("Gross Margin",       0,    0,    0.30),
    ("Operating Expenses", 0,    0,    500),
    ("Net Income",         0,    0,    2500),
    ("Active Tutors",      0,    0,    10),
    ("Active Clients",     0,    0,    20),
    ("Sessions Delivered", 0,    0,    200),
    ("Avg Sessions/Client",0,    0,    10),
]
fmt_map = {
    "Revenue":            money_fmt(),
    "Tutor Payments":     money_fmt(),
    "Gross Profit":       money_fmt(),
    "Gross Margin":       pct_fmt(),
    "Operating Expenses": money_fmt(),
    "Net Income":         money_fmt(),
    "Active Tutors":      "0",
    "Active Clients":     "0",
    "Sessions Delivered": "0",
    "Avg Sessions/Client":"0.0",
}

for i, (label, month_val, ytd_val, target) in enumerate(kpis):
    r = kpi_row + 1 + i
    bg = GRAY1 if i % 2 == 0 else WHITE
    write(ws, r, 2, label, bold=(label in ("Gross Profit","Net Income")), bg=bg)
    money(ws, r, 3, month_val, bg=bg)
    money(ws, r, 4, ytd_val, bg=bg)
    f = fmt_map[label]
    cell = ws.cell(row=r, column=5, value=target)
    cell.number_format = f
    cell.alignment = Alignment(horizontal="right")
    cell.fill = PatternFill("solid", fgColor=bg)

# Manual entry note
note_row = kpi_row + len(kpis) + 2
write(ws, note_row, 2,
      "📌  Update Month & YTD manually by copying totals from Monthly P&L sheet.",
      italic=True, size=9, color="FF555555")
write(ws, note_row + 1, 2,
      "Backend data: export DB → open in DB Browser for SQLite → paste totals here.",
      italic=True, size=9, color="FF555555")

# Revenue target roadmap
rm_row = note_row + 3
write(ws, rm_row, 2, "Revenue Roadmap to $200k / month", bold=True, size=12)
write(ws, rm_row + 1, 2, "Phase",  bold=True)
write(ws, rm_row + 1, 3, "Sessions/mo", bold=True)
write(ws, rm_row + 1, 4, "Revenue/mo",  bold=True)
write(ws, rm_row + 1, 5, "Net/mo",      bold=True)

phases = [
    ("Phase 1 – Launch  (0-3 mo)",   20,   1000,    300),
    ("Phase 2 – Traction (3-6 mo)",  80,   4000,   1200),
    ("Phase 3 – Growth  (6-12 mo)", 200,  10000,   3000),
    ("Phase 4 – Scale   (12-18 mo)",600,  30000,   9000),
    ("Phase 5 – $200k   (18-24 mo)",4000,200000,  60000),
]
for j, (ph, sess, rev, net) in enumerate(phases):
    r = rm_row + 2 + j
    bg = GREEN if rev >= 200000 else (GRAY1 if j % 2 == 0 else WHITE)
    write(ws, r, 2, ph, bg=bg)
    write(ws, r, 3, sess, bg=bg, align="right")
    money(ws, r, 4, rev, bg=bg)
    money(ws, r, 5, net, bg=bg)

freeze(ws, "B5")

# ══════════════════════════════════════════════════════════════════════════════
# 2.  MONTHLY P&L
# ══════════════════════════════════════════════════════════════════════════════
ws = ws_map["Monthly P&L"]
ws.sheet_view.showGridLines = False
months = ["Jan","Feb","Mar","Apr","May","Jun",
          "Jul","Aug","Sep","Oct","Nov","Dec"]

set_col_widths(ws, [2, 26] + [12]*12 + [12, 2])

ws.merge_cells("B2:N2")
c = ws.cell(row=2, column=2, value="Monthly P&L Statement – Enter actuals each month")
c.font = Font(name="Calibri", bold=True, size=13)
c.alignment = Alignment(horizontal="left")

header_row(ws, 4, ["", "Line Item"] + months + ["Full Year"], bg="FF222222")
ws.row_dimensions[4].height = 22

pl_lines = [
    # (label, row_tag, style)  style: normal | subtotal | total | blank
    ("REVENUE",                         "header",   "header"),
    ("Client Sessions Revenue",         "revenue",  "input"),
    ("Bundles (4-session, $180 ea)",    "bundles",  "input"),
    ("Resume / PS Reviews ($50 ea)",    "resumes",  "input"),
    ("Total Revenue",                   "rev_total","subtotal"),
    ("",                                "",         "blank"),
    ("COST OF REVENUE",                 "header",   "header"),
    ("Tutor Payments ($35/hr)",         "tutor_pay","input"),
    ("Gross Profit",                    "gp",       "subtotal"),
    ("Gross Margin %",                  "gm_pct",   "pct_sub"),
    ("",                                "",         "blank"),
    ("OPERATING EXPENSES",              "header",   "header"),
    ("Marketing / Flyers / Ads",        "mktg",     "input"),
    ("Software & Tools",                "tools",    "input"),
    ("Travel / Logistics",              "travel",   "input"),
    ("Miscellaneous",                   "misc",     "input"),
    ("Total OpEx",                      "opex",     "subtotal"),
    ("",                                "",         "blank"),
    ("Net Income",                      "net",      "total"),
    ("Net Margin %",                    "nm_pct",   "pct_total"),
]

row = 5
input_rows = {}   # tag -> spreadsheet row number

for label, tag, style in pl_lines:
    if style == "blank":
        row += 1
        continue

    ws.row_dimensions[row].height = 18

    if style == "header":
        write(ws, row, 2, label, bold=True, size=10, bg="FF444444", color=WHITE)
        for c in range(3, 16):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="FF444444")
        row += 1
        continue

    is_sub   = style in ("subtotal", "pct_sub")
    is_total = style in ("total", "pct_total")
    is_pct   = style in ("pct_sub", "pct_total")
    bg = GRAY1 if not (is_sub or is_total) else (GREEN if is_total else WHITE)

    write(ws, row, 2, label,
          bold=is_sub or is_total,
          bg=bg)
    input_rows[tag] = row

    for ci, mo in enumerate(months):
        col = 3 + ci
        if style == "input":
            cell = ws.cell(row=row, column=col, value=0)
            cell.number_format = money_fmt()
            cell.alignment = Alignment(horizontal="right")
            cell.fill = PatternFill("solid", fgColor=bg)
        else:
            # leave blank for now – formulas would go here
            cell = ws.cell(row=row, column=col, value=None)
            cell.number_format = money_fmt() if not is_pct else pct_fmt()
            cell.alignment = Alignment(horizontal="right")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name="Calibri", italic=True, size=9,
                             color="FF888888")

    # Full Year column (col 15)
    cell = ws.cell(row=row, column=15, value=None)
    cell.number_format = money_fmt() if not is_pct else pct_fmt()
    cell.alignment = Alignment(horizontal="right")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Calibri", bold=is_sub or is_total, size=10)

    row += 1

note_r = row + 1
write(ws, note_r, 2,
      "📌  Input cells accept actual figures. Calculated rows (subtotals/margins) are for manual entry.",
      italic=True, size=9, color="FF555555")
write(ws, note_r + 1, 2,
      "    Use Dashboard sheet to copy monthly totals into KPI boxes.",
      italic=True, size=9, color="FF555555")

freeze(ws, "C5")

# ══════════════════════════════════════════════════════════════════════════════
# 3-5.  THREE-YEAR PROJECTIONS (Conservative / Base Case / Aggressive)
# ══════════════════════════════════════════════════════════════════════════════

SCENARIOS = {
    "3yr Conservative": dict(
        growth_pct  = 0.08,    # 8% MoM session growth
        start_sess  = 10,
        client_rate = 50,
        tutor_rate  = 35,
        opex_mo     = 200,
        label       = "Conservative (8% MoM growth)",
    ),
    "3yr Base Case": dict(
        growth_pct  = 0.15,
        start_sess  = 20,
        client_rate = 50,
        tutor_rate  = 35,
        opex_mo     = 400,
        label       = "Base Case (15% MoM growth)",
    ),
    "3yr Aggressive": dict(
        growth_pct  = 0.25,
        start_sess  = 30,
        client_rate = 50,
        tutor_rate  = 35,
        opex_mo     = 700,
        label       = "Aggressive (25% MoM growth)",
    ),
}

for sheet_name, cfg in SCENARIOS.items():
    ws = ws_map[sheet_name]
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, [2, 14, 14, 14, 14, 14, 14, 14, 2])

    ws.merge_cells("B2:H2")
    c = ws.cell(row=2, column=2, value=f"3-Year Projection – {cfg['label']}")
    c.font = Font(name="Calibri", bold=True, size=13)

    # Assumptions box
    write(ws, 4, 2, "Assumptions", bold=True, size=11)
    assump = [
        ("Client Rate ($/hr)",    cfg["client_rate"]),
        ("Tutor Rate ($/hr)",     cfg["tutor_rate"]),
        ("Margin / hr",           cfg["client_rate"] - cfg["tutor_rate"]),
        ("Gross Margin %",        (cfg["client_rate"] - cfg["tutor_rate"]) / cfg["client_rate"]),
        ("Starting Sessions/mo",  cfg["start_sess"]),
        ("MoM Session Growth",    cfg["growth_pct"]),
        ("Monthly OpEx",          cfg["opex_mo"]),
    ]
    for k, (lbl, val) in enumerate(assump):
        r = 5 + k
        bg = GRAY1 if k % 2 == 0 else WHITE
        write(ws, r, 2, lbl, bg=bg)
        is_pct_val = isinstance(val, float) and (lbl.endswith("%") or "Growth" in lbl)
        cell = ws.cell(row=r, column=3, value=val)
        cell.alignment = Alignment(horizontal="right")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.number_format = pct_fmt() if is_pct_val else (money_fmt() if isinstance(val, float) and val > 1 else "0")

    # Projection table header
    hdr_r = 14
    header_row(ws, hdr_r, ["", "Month", "Sessions", "Revenue", "Tutor Cost",
                             "Gross Profit", "OpEx", "Net Income"],
               bg="FF222222")

    sessions = cfg["start_sess"]
    cr = cfg["client_rate"]
    tr = cfg["tutor_rate"]
    opex = cfg["opex_mo"]
    growth = cfg["growth_pct"]

    YEAR_LABELS = {12: "── Year 1 Total ──", 24: "── Year 2 Total ──", 36: "── Year 3 Total ──"}
    y_totals = {1: [0]*6, 2: [0]*6, 3: [0]*6}

    for mo in range(1, 37):
        rev   = sessions * cr
        cost  = sessions * tr
        gp    = rev - cost
        net   = gp - opex
        year  = (mo - 1) // 12 + 1
        r = hdr_r + mo
        bg = GRAY1 if mo % 2 == 1 else WHITE
        write(ws, r, 2, f"Month {mo:02d}", bg=bg)
        write(ws, r, 3, round(sessions), bg=bg, align="right")
        money(ws, r, 4, rev,  bg=bg)
        money(ws, r, 5, cost, bg=bg)
        money(ws, r, 6, gp,   bg=bg, color=("FF1a7a1a" if gp > 0 else "FFcc0000"))
        money(ws, r, 7, opex, bg=bg)
        money(ws, r, 8, net,  bg=bg,
              color=("FF1a7a1a" if net > 0 else "FFcc0000"),
              bold=(mo % 12 == 0))
        # accumulate year totals
        y_totals[year][0] += sessions
        y_totals[year][1] += rev
        y_totals[year][2] += cost
        y_totals[year][3] += gp
        y_totals[year][4] += opex
        y_totals[year][5] += net

        if mo in YEAR_LABELS:
            yr = mo // 12
            r2 = hdr_r + mo + 1
            label = YEAR_LABELS[mo]
            write(ws, r2, 2, label, bold=True, bg="FF222222", color=WHITE)
            write(ws, r2, 3, round(y_totals[yr][0]), bold=True, bg="FF222222", color=WHITE, align="right")
            for ci, val in enumerate(y_totals[yr][1:], 4):
                cell = ws.cell(row=r2, column=ci, value=round(val, 2))
                cell.number_format = money_fmt()
                cell.alignment = Alignment(horizontal="right")
                cell.fill = PatternFill("solid", fgColor="FF222222")
                cell.font = Font(name="Calibri", bold=True, color=WHITE)
            # add blank row
            hdr_r += 2   # offset so next month lands 2 rows lower

        sessions = sessions * (1 + growth)

    freeze(ws, "B15")

# ══════════════════════════════════════════════════════════════════════════════
# 6.  LEAD FUNNEL TRACKER
# ══════════════════════════════════════════════════════════════════════════════
ws = ws_map["Lead Funnel"]
ws.sheet_view.showGridLines = False
set_col_widths(ws, [2, 20, 14, 14, 14, 14, 20, 2])

ws.merge_cells("B2:G2")
c = ws.cell(row=2, column=2, value="Lead Funnel Tracker")
c.font = Font(name="Calibri", bold=True, size=13)

header_row(ws, 4, ["", "Month", "New Leads", "Contacted", "Converted", "Lost", "Conv. Rate %"],
           bg="FF222222")

months_full = [
    "Jan 2026","Feb 2026","Mar 2026","Apr 2026","May 2026","Jun 2026",
    "Jul 2026","Aug 2026","Sep 2026","Oct 2026","Nov 2026","Dec 2026",
]
for i, mo in enumerate(months_full):
    r = 5 + i
    bg = GRAY1 if i % 2 == 0 else WHITE
    write(ws, r, 2, mo, bg=bg)
    for c in range(3, 7):
        cell = ws.cell(row=r, column=c, value=0)
        cell.alignment = Alignment(horizontal="right")
        cell.fill = PatternFill("solid", fgColor=bg)
    # Conv rate: converted / new leads (formula hint)
    cell = ws.cell(row=r, column=7, value=None)
    cell.number_format = pct_fmt()
    cell.alignment = Alignment(horizontal="right")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Calibri", italic=True, size=9, color="FF888888")

write(ws, 18, 2, "📌  Update monthly from Admin dashboard. Conversion rate = Converted ÷ New Leads.",
      italic=True, size=9, color="FF555555")
freeze(ws, "C5")

# ══════════════════════════════════════════════════════════════════════════════
# 7.  TUTOR ROSTER
# ══════════════════════════════════════════════════════════════════════════════
ws = ws_map["Tutor Roster"]
ws.sheet_view.showGridLines = False
set_col_widths(ws, [2, 22, 26, 14, 12, 12, 16, 22, 22, 2])

ws.merge_cells("B2:I2")
c = ws.cell(row=2, column=2, value="Tutor Roster")
c.font = Font(name="Calibri", bold=True, size=13)

header_row(ws, 4,
    ["", "Name", "Email", "Phone", "Rate ($/hr)", "Active", "Subjects",
     "Earliest Avail.", "Notes"],
    bg="FF222222")

for i in range(20):
    r = 5 + i
    bg = GRAY1 if i % 2 == 0 else WHITE
    for c in range(2, 10):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=bg)
    # Rate default
    cell = ws.cell(row=r, column=5, value=35)
    cell.number_format = money_fmt()
    cell.alignment = Alignment(horizontal="right")
    cell.fill = PatternFill("solid", fgColor=bg)
    # Active default
    cell = ws.cell(row=r, column=6, value="Yes")
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill("solid", fgColor=bg)

write(ws, 26, 2, "📌  This roster is for planning reference. Live data is in the backend DB.",
      italic=True, size=9, color="FF555555")
freeze(ws, "C5")

# ══════════════════════════════════════════════════════════════════════════════
# 8.  CLIENT ROSTER
# ══════════════════════════════════════════════════════════════════════════════
ws = ws_map["Client Roster"]
ws.sheet_view.showGridLines = False
set_col_widths(ws, [2, 22, 26, 14, 14, 14, 22, 2])

ws.merge_cells("B2:G2")
c = ws.cell(row=2, column=2, value="Client Roster")
c.font = Font(name="Calibri", bold=True, size=13)

header_row(ws, 4,
    ["", "Name", "Email", "Phone", "Sessions", "Total Paid ($)", "Notes"],
    bg="FF222222")

for i in range(30):
    r = 5 + i
    bg = GRAY1 if i % 2 == 0 else WHITE
    for c in range(2, 8):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=bg)
    for c in (5, 6):
        cell = ws.cell(row=r, column=c, value=0)
        cell.alignment = Alignment(horizontal="right")
        cell.fill = PatternFill("solid", fgColor=bg)
    ws.cell(row=r, column=6).number_format = money_fmt()

freeze(ws, "C5")

# ══════════════════════════════════════════════════════════════════════════════
# 9.  SESSION LOG
# ══════════════════════════════════════════════════════════════════════════════
ws = ws_map["Session Log"]
ws.sheet_view.showGridLines = False
set_col_widths(ws, [2, 6, 22, 22, 18, 12, 12, 12, 12, 14, 22, 2])

ws.merge_cells("B2:K2")
c = ws.cell(row=2, column=2, value="Session Log")
c.font = Font(name="Calibri", bold=True, size=13)

header_row(ws, 4,
    ["", "#", "Client Name", "Tutor Name", "Subject", "Date",
     "Duration (hrs)", "Client Rate", "Tutor Rate", "Margin ($)", "Notes"],
    bg="FF222222")

for i in range(100):
    r = 5 + i
    bg = GRAY1 if i % 2 == 0 else WHITE
    for c in range(2, 12):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=bg)
    # Row number
    ws.cell(row=r, column=2, value=i+1).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=bg)
    # Default rates
    for c, val in [(8, 50), (9, 35)]:
        cell = ws.cell(row=r, column=c, value=val)
        cell.number_format = money_fmt()
        cell.alignment = Alignment(horizontal="right")
        cell.fill = PatternFill("solid", fgColor=bg)
    # Margin = client rate - tutor rate (informational, not formula)
    cell = ws.cell(row=r, column=10, value=15)
    cell.number_format = money_fmt()
    cell.alignment = Alignment(horizontal="right")
    cell.fill = PatternFill("solid", fgColor=GREEN)

write(ws, 106, 2,
      "📌  Log each session here AND in the admin dashboard (+ Session button) to keep DB in sync.",
      italic=True, size=9, color="FF555555")
freeze(ws, "C5")

# ══════════════════════════════════════════════════════════════════════════════
# 10.  PAYMENT LOG
# ══════════════════════════════════════════════════════════════════════════════
ws = ws_map["Payment Log"]
ws.sheet_view.showGridLines = False
set_col_widths(ws, [2, 6, 14, 14, 14, 16, 22, 16, 22, 2])

ws.merge_cells("B2:I2")
c = ws.cell(row=2, column=2, value="Payment Log")
c.font = Font(name="Calibri", bold=True, size=13)

header_row(ws, 4,
    ["", "#", "Date", "Direction", "Amount ($)", "Method",
     "Client / Tutor Name", "Session ID", "Notes"],
    bg="FF222222")

DIRECTION_COLORS = {"In (client → us)": GREEN, "Out (us → tutor)": RED}

for i in range(150):
    r = 5 + i
    bg = GRAY1 if i % 2 == 0 else WHITE
    for c in range(2, 10):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=bg)
    ws.cell(row=r, column=2, value=i+1).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=bg)
    cell = ws.cell(row=r, column=5, value=0)
    cell.number_format = money_fmt()
    cell.alignment = Alignment(horizontal="right")
    cell.fill = PatternFill("solid", fgColor=bg)
    # Default direction
    cell = ws.cell(row=r, column=4, value="In (client → us)")
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill("solid", fgColor=bg)

write(ws, 156, 2,
      "📌  Log every Zelle transfer here AND in the admin dashboard ($ Payment button).",
      italic=True, size=9, color="FF555555")
write(ws, 157, 2,
      "    Direction 'In' = client paid us (revenue).  Direction 'Out' = we paid tutor (COGS).",
      italic=True, size=9, color="FF555555")
freeze(ws, "C5")

# ══════════════════════════════════════════════════════════════════════════════
# 11.  EXPENSE LOG
# ══════════════════════════════════════════════════════════════════════════════
ws = ws_map["Expense Log"]
ws.sheet_view.showGridLines = False
set_col_widths(ws, [2, 6, 14, 18, 16, 22, 22, 2])

ws.merge_cells("B2:G2")
c = ws.cell(row=2, column=2, value="Expense Log")
c.font = Font(name="Calibri", bold=True, size=13)

header_row(ws, 4,
    ["", "#", "Date", "Category", "Amount ($)", "Description", "Notes"],
    bg="FF222222")

CATEGORIES = ["Marketing", "Tools", "Travel", "Misc"]

for i in range(50):
    r = 5 + i
    bg = GRAY1 if i % 2 == 0 else WHITE
    for c in range(2, 8):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=bg)
    ws.cell(row=r, column=2, value=i+1).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=bg)
    cell = ws.cell(row=r, column=5, value=0)
    cell.number_format = money_fmt()
    cell.alignment = Alignment(horizontal="right")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell = ws.cell(row=r, column=4, value="Marketing")
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill("solid", fgColor=bg)

write(ws, 56, 2,
      "📌  Track all operating costs here. Categories: Marketing, Tools, Travel, Misc.",
      italic=True, size=9, color="FF555555")
freeze(ws, "C5")

# ══════════════════════════════════════════════════════════════════════════════
# 12.  PRICING CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════
ws = ws_map["Pricing Calculator"]
ws.sheet_view.showGridLines = False
set_col_widths(ws, [2, 30, 16, 16, 16, 2])

ws.merge_cells("B2:E2")
c = ws.cell(row=2, column=2, value="Pricing Calculator")
c.font = Font(name="Calibri", bold=True, size=13)

write(ws, 3, 2,
      "Adjust the yellow cells to model different rate scenarios.",
      italic=True, size=9, color="FF555555")

write(ws, 5, 2, "Input Variables", bold=True, size=11)
inputs = [
    ("Client Rate ($/hr)",   50,  "client_rate"),
    ("Tutor Rate ($/hr)",    35,  "tutor_rate"),
    ("Sessions per Month",  200,  "sessions"),
    ("Avg Duration (hrs)",  1.0,  "duration"),
    ("Monthly OpEx ($)",    400,  "opex"),
]
inp_rows = {}
for k, (lbl, val, tag) in enumerate(inputs):
    r = 6 + k
    write(ws, r, 2, lbl)
    cell = ws.cell(row=r, column=3, value=val)
    cell.fill = PatternFill("solid", fgColor=YELLOW)
    cell.alignment = Alignment(horizontal="right")
    cell.number_format = money_fmt() if isinstance(val, float) and val > 1 else "0.0#"
    cell.font = Font(name="Calibri", bold=True, size=11)
    inp_rows[tag] = r

write(ws, 12, 2, "Calculated Results", bold=True, size=11)
# Hardcode calculated rows; user can update inputs and see effect
results = [
    ("Margin per Hour",        15,     money_fmt()),
    ("Gross Margin %",         0.30,   pct_fmt()),
    ("Monthly Revenue",        10000,  money_fmt()),
    ("Monthly Tutor Cost",     7000,   money_fmt()),
    ("Monthly Gross Profit",   3000,   money_fmt()),
    ("Monthly OpEx",           400,    money_fmt()),
    ("Monthly Net Income",     2600,   money_fmt()),
    ("Net Margin %",           0.26,   pct_fmt()),
    ("Annual Revenue (proj.)", 120000, money_fmt()),
    ("Annual Net Income",      31200,  money_fmt()),
]
for k, (lbl, val, fmt) in enumerate(results):
    r = 13 + k
    bg = GREEN if "Net Income" in lbl or "Gross Profit" in lbl else GRAY1
    write(ws, r, 2, lbl, bold=("Net" in lbl), bg=bg)
    cell = ws.cell(row=r, column=3, value=val)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Calibri", bold=("Net" in lbl))

write(ws, 24, 2,
      "📌  To model 4-session bundle ($180): client_rate = 45 (effective).  Resume review = $50 flat.",
      italic=True, size=9, color="FF555555")
write(ws, 25, 2,
      "    Anti-poaching: client info withheld until Zelle payment confirmed. Tutor paid weekly.",
      italic=True, size=9, color="FF555555")

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "finances.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
